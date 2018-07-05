# Import Files
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows 
exportPath = r'C:\Users\Rohan\Desktop\PythonExport.xlsx'

# Creating a dictionary
dict1={}

dict1['column1']=['r_n_1','r_n_1','r_n_1','r_n_1','r_n_10', 'r_n_10']
dict1['column2']=['r_s_1','r_s_1','r_s_1','r_s_1','r_s_7', 'r_s_9']
dict1['column3']=['r_n_2','r_n_4','r_n_6','r_n_6','r_n_11', 'r_n_11']
dict1['column4']=['r_s_3','r_s_4','r_s_5','r_s_6','r_s_8', 'r_s_10']
dict1['column5']=['r_n_3','r_n_5','r_n_7','r_n_9','r_n_12', 'r_n_13']

#Converting the dictionary to a Data Frame
df = pd.DataFrame(dict1)

print(df)

"""
Output:

  column1 column2 column3 column4 column5
0   r_n_1   r_s_1   r_n_2   r_s_3   r_n_3
1   r_n_1   r_s_1   r_n_4   r_s_4   r_n_5
2   r_n_1   r_s_1   r_n_6   r_s_5   r_n_7
3   r_n_1   r_s_1   r_n_6   r_s_6   r_n_9
4  r_n_10   r_s_7  r_n_11   r_s_8  r_n_12
5  r_n_10   r_s_9  r_n_11  r_s_10  r_n_13
"""

# To identify the repeating values
for i in df.columns:
    for j in range(len(df[i])):        
        for k in range(j+1,len(df[i])):
            if df[i][j]== df[i][k]:                
                df[i][k]='-' 
print("")
print(df)

"""
Output:

  column1 column2 column3 column4 column5
0   r_n_1   r_s_1   r_n_2   r_s_3   r_n_3
1       -       -   r_n_4   r_s_4   r_n_5
2       -       -   r_n_6   r_s_5   r_n_7
3       -       -       -   r_s_6   r_n_9
4  r_n_10   r_s_7  r_n_11   r_s_8  r_n_12
5       -   r_s_9       -  r_s_10  r_n_13
"""

# Merge the similar cells
wb= Workbook()
ws=wb.active

rowInd=1
colInd=1

print("")

for i in df.columns:
    for j in range(0,len(df[i])):
        if(df[i][j]!='-'):
            ws.cell(row=rowInd,column=colInd,value=df[i][j])            
        else:
            count=0
            for l in range(j+1,len(df[i])):
                count+=1
                if df[i][l]!='-':
                    count-=1
                    break
            ws.merge_cells(str(str(colList[colInd]+str(rowInd-1))+":"+str(colList[colInd]+str(rowInd+count))))
        rowInd+=1
            
    colInd+=1
    rowInd=1  
    
wb.save(exportPath)
